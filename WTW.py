import openpyxl as op
def cumulative(input):
    import pandas as pd
    df=pd.read_csv('input.csv')
    df.to_excel('output.xlsx','Sheet1')
    workbook=op.load_workbook(filename='output.xlsx')
    workbook.save('output.xlsx')
    sheet=workbook['Sheet1']
    
    sheet.sort_values(by='product') # here I am sorting the product column of the table, so that rows that have the same products are on top of each other
     # below is a for loop where I am trying to create multiple excel worksheets in the file 'output.xlsx' where each work sheet is for a unique Product.
    for i in range(1,sheet.max_row+1):
        if sheet.cell(i,1) != sheet.cell(i+1,1): # here if the product in the current row is different to the product (new product) 
            # in the next row, then I will create a new excel work sheet and paste the row that contains the new product onto the new worksheet.
           workbook.create_sheet(f"Sheet{i+1}")
           Sheet=workbook[f"Sheet{i+1}"]
           workbook.save('output.xlsx')
           for j in range(1,sheet.max_col+1):
               Sheet.cell(i,j).value=sheet.cell(i,j).value
        else:
            workbook.create_sheet(f"Sheet{i}")
            Sheet=workbook[f"Sheet{i}"]
            for k in range(1,sheet.max_col+1):
               Sheet.cell(i,k).value=sheet.cell(i,k).value
            workbook.save('output.xlsx')
    workbook.worksheets()
    ## Now i am planning to access each worksheet of the excel file "output.xlsx", then sort the orgin year column in ascending order, then create a new column that contains the
    # # cumalative development values, then create a seperate worksheet such that in 
    # # the 1st row it has the earliest origin year and the maximum number of development years
   # then in the consecutive rows, thier will be the pri=duct name in the first column, then the cumulative of the claims paid at each development year, then I woul convert
   # this worksheet into a csv file, hence this will be the output.